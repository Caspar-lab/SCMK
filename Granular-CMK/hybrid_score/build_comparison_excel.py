"""
build_comparison_excel.py — 汇总所有算法在各数据集上的最优 AUC 到一份 Excel
==========================================================================
列：dataset, group, N, D, anomaly | scatter(ours) | OCSVM | DeepSVDD | <Experimental_results 全部算法>
  · scatter   : hybrid_best.csv (max_ensemble 最优)
  · OCSVM     : 现场 RBF 半监督 OCSVM, nu 网格取最优(与 scatter 同预处理同协议)
  · DeepSVDD  : PyOD 版 deepsvdd_best.csv
  · 其它      : Experimental_results/{alg}_results/{ds}/{ds}_{alg}.mat 的 opt_out_scores[:,0]
格式：每行最优 AUC 高亮、scatter 列着色、冻结首行+前列。
输出：C:/OD/Shihao/5/result/algorithm_comparison.xlsx
"""
import os, sys
import numpy as np, pandas as pd, scipy.io as sio
from sklearn.metrics import roc_auc_score
from sklearn.svm import OneClassSVM
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from run_hybrid_score import load_data, NU_CANDIDATES

ER  = r'C:\OD\Shihao\Experimental_results'
DS  = r'C:\OD\Shihao\datasets'
HB  = r'C:\OD\Shihao\5\result\hybrid_score\hybrid_best.csv'
DSV = r'C:\OD\Shihao\5\Granular-CMK\hybrid_score\deepsvdd\deepsvdd_best.csv'
UNSUP = r'C:\OD\Shihao\5\result\scatter_unsup\scatter_unsup_best.csv'
OUT = r'C:\OD\Shihao\5\result\algorithm_comparison.xlsx'

SKIP_ALG = {'DeepSVDD', 'OCSVM'}          # 用我们自己的 PyOD-DeepSVDD / 现场-OCSVM 版本


def label_of(stem):
    p = os.path.join(DS, stem + '.mat')
    if not os.path.exists(p):
        return None
    d = sio.loadmat(p)
    return (d['trandata'][:, -1] != 0).astype(int) if 'trandata' in d else None


def find_main_mat(alg, stem):
    folder = os.path.join(ER, f'{alg}_results', stem)
    if not os.path.isdir(folder):
        return None
    exact = os.path.join(folder, f'{stem}_{alg}.mat')
    if os.path.exists(exact):
        return exact
    c = [f for f in os.listdir(folder) if f.endswith(f'{stem}_{alg}.mat')]
    if c:
        return os.path.join(folder, sorted(c, key=len)[0])
    c = [f for f in os.listdir(folder)
         if f.endswith(f'_{alg}.mat') and '_k-' not in f and '_run' not in f]
    return os.path.join(folder, sorted(c, key=len)[0]) if c else None


def alg_auc(alg, stem, y):
    p = find_main_mat(alg, stem)
    if p is None:
        return np.nan
    try:
        s = np.asarray(sio.loadmat(p)['opt_out_scores'])[:, 0].ravel()
        return roc_auc_score(y, s) if len(s) == len(y) else np.nan
    except Exception:
        return np.nan


def ocsvm_auc(X, y):
    best = -1.0
    for nu in NU_CANDIDATES:
        try:
            clf = OneClassSVM(kernel='rbf', gamma='scale', nu=nu).fit(X[y == 0])
            best = max(best, roc_auc_score(y, -clf.decision_function(X)))
        except Exception:
            pass
    return best if best > 0 else np.nan


if __name__ == '__main__':
    try: sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except Exception: pass

    hb = pd.read_csv(HB)
    scatter_map = dict(zip(hb['dataset'], hb['best_auc']))
    group_map   = dict(zip(hb['dataset'], hb['group']))
    datasets = list(hb['dataset'])

    dsv = pd.read_csv(DSV)
    dsvdd_map = dict(zip(dsv['dataset'], dsv['best_auc']))
    unsup_map = (dict(zip(pd.read_csv(UNSUP)['dataset'], pd.read_csv(UNSUP)['best_auc']))
                 if os.path.exists(UNSUP) else {})

    # Experimental_results 全部算法（去重复/特殊版本）
    exp_algs = sorted({f[:-len('_results')] for f in os.listdir(ER)
                       if f.endswith('_results') and os.path.isdir(os.path.join(ER, f))}
                      - SKIP_ALG)
    print(f'数据集={len(datasets)}  Experimental_results 算法={len(exp_algs)}')

    rows = []
    for i, stem in enumerate(datasets, 1):
        y = label_of(stem)
        if y is None:
            print(f'  [skip] {stem}: 无标签')
            continue
        X, yy, meta = load_data(os.path.join(DS, stem + '.mat'))
        rec = {'dataset': stem, 'group': group_map.get(stem, ''),
               'N': meta['N'], 'D': X.shape[1],
               'anomaly': round(meta['anomaly_rate'], 4),
               'scatter': round(float(scatter_map[stem]), 4),
               'scatter_unsup': (round(float(unsup_map[stem]), 4)
                                 if stem in unsup_map else np.nan),
               'OCSVM': round(ocsvm_auc(X, yy), 4),
               'DeepSVDD': (round(float(dsvdd_map[stem]), 4)
                            if stem in dsvdd_map else np.nan)}
        for alg in exp_algs:
            a = alg_auc(alg, stem, y)
            rec[alg] = round(a, 4) if not np.isnan(a) else np.nan
        rows.append(rec)
        print(f'  [{i}/{len(datasets)}] {stem}  scatter={rec["scatter"]} OCSVM={rec["OCSVM"]} DeepSVDD={rec["DeepSVDD"]}')

    df = pd.DataFrame(rows)
    meta_cols = ['dataset', 'group', 'N', 'D', 'anomaly']
    alg_cols  = ['scatter', 'scatter_unsup', 'OCSVM', 'DeepSVDD'] + exp_algs
    df = df[meta_cols + alg_cols].sort_values(['group', 'D', 'dataset']).reset_index(drop=True)
    df.to_excel(OUT, index=False, sheet_name='AUC')

    # ── openpyxl 格式化 ──
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    wb = load_workbook(OUT); ws = wb['AUC']
    hdr_fill = PatternFill('solid', fgColor='305496')
    scat_fill = PatternFill('solid', fgColor='DDEBF7')   # scatter 列浅蓝
    best_fill = PatternFill('solid', fgColor='FFF2CC')   # 每行最优浅黄
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c); cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = hdr_fill; cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'F2'                               # 冻结前5列meta + 表头
    first_alg = len(meta_cols) + 1                       # scatter 所在列号(1-based)
    scat_col = first_alg
    for r in range(2, ws.max_row + 1):
        vals = []
        for c in range(first_alg, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                vals.append((v, c))
        ws.cell(r, scat_col).fill = scat_fill            # scatter 列着色
        ws.cell(r, scat_col + 1).fill = scat_fill        # scatter_unsup 列着色
        if vals:
            mx = max(vals)[0]
            for v, c in vals:
                if abs(v - mx) < 1e-9:
                    ws.cell(r, c).fill = best_fill        # 每行最优高亮
    ws.column_dimensions['A'].width = 34
    wb.save(OUT)
    print(f'\n完成。{len(df)} 行 × {len(alg_cols)} 算法  -> {OUT}')
