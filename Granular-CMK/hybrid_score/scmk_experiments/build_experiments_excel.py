"""
build_experiments_excel.py — compile every SCMK experiment into one workbook.
================================================================================
Gathers, into result/scmk_experiments/SCMK_experiments_summary.xlsx:

  * the ORIGINAL SCMK hyperparameter study (dual-signal fused AUC on the
    semi-supervised split), across latent_dim x lambda_scatter x split_seed
    -> result/hybrid_score_semi/hybrid_semi_{all,mean,best}.csv
  * the NEW loss-architecture experiments added on top of SCMK (fixed dim=64,
    lambda=100, seed=2, 19 datasets = 20 minus pageblocks):
       - temperature sweep, EMBEDDING bandwidth   (temp_sweep_emb_results.csv)
       - temperature sweep, RAW bandwidth         (temp_sweep_results.csv)
       - 2x2 factorial  bandwidth x temperature   (factorial_all_results.csv)
       - cosine NT-Xent vs kernel-logit loss      (cosine_vs_kernel_results.csv)

plus an Overview sheet, a parameter-effect pivot for the original grid, and a
head-to-head summary of every new loss variant vs the original SCMK loss.

Run:  conda run -n torch311 python build_experiments_excel.py
"""
import os
import numpy as np
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(os.path.dirname(HERE)))      # .../5
RES = os.path.join(ROOT, 'result', 'scmk_experiments')             # new-experiment outputs
SEMI = os.path.join(ROOT, 'result', 'hybrid_score_semi')           # original SCMK grid
OUT = os.path.join(RES, 'SCMK_experiments_summary.xlsx')
TAUS = [1.0, 0.7, 0.5, 0.3, 0.2, 0.1, 0.05]


def _read(path):
    return pd.read_csv(path) if os.path.exists(path) else None


# ── original SCMK hyperparameter study ───────────────────────────────────────
grid_all = _read(os.path.join(SEMI, 'hybrid_semi_all.csv'))    # ds x dim x lam x seed
grid_mean = _read(os.path.join(SEMI, 'hybrid_semi_mean.csv'))  # mean +/- std over seeds
grid_best = _read(os.path.join(SEMI, 'hybrid_semi_best.csv'))  # best (dim,lam) per dataset


def param_effect(df):
    """Marginal effect of each hyperparameter on SCMK AUC (mean over everything else)."""
    by_dim = df.groupby('latent_dim')['auc'].agg(['mean', 'std', 'count']).reset_index()
    by_lam = df.groupby('lambda_scatter')['auc'].agg(['mean', 'std', 'count']).reset_index()
    by_seed = df.groupby('split_seed')['auc'].agg(['mean', 'std', 'count']).reset_index()
    pivot = df.pivot_table(index='latent_dim', columns='lambda_scatter', values='auc', aggfunc='mean')
    return by_dim, by_lam, by_seed, pivot


# ── new loss-architecture experiments ────────────────────────────────────────
emb = _read(os.path.join(RES, 'temp_sweep_emb_results.csv'))        # embedding bw tau sweep
raw = _read(os.path.join(RES, 'temp_sweep_results.csv'))            # raw bw tau sweep
fac = _read(os.path.join(RES, 'factorial_all_results.csv'))         # 2x2 factorial
cos = _read(os.path.join(RES, 'cosine_vs_kernel_results.csv'))      # cosine vs kernel loss


def annotate_best(df, metric):
    """Add best_t / best_<metric> / gain_vs_SCMK columns over the tau grid."""
    df = df.copy()
    cols = [f'{metric}_t{t}' for t in TAUS if f'{metric}_t{t}' in df.columns]
    vals = df[cols].values
    bi = vals.argmax(axis=1)
    df[f'best_t_{metric}'] = [TAUS[i] for i in bi]
    df[f'best_{metric}'] = vals[np.arange(len(df)), bi]
    return df


def build_summary():
    """One row per dataset: original SCMK loss vs each new loss variant (fused AUC)."""
    rows = []
    for _, r in emb.iterrows():
        d = r['dataset']
        rec = {'dataset': d, 'raw_RBF': r['raw_rbf'], 'SCMK_orig_fused': r['scmk_fused']}
        # embedding-bw temperature sweep (best tau)
        vals = [r[f'fused_t{t}'] for t in TAUS]
        bi = int(np.argmax(vals))
        rec['embBW_bestT'] = TAUS[bi]; rec['embBW_fused'] = vals[bi]
        # raw-bw temperature sweep (best tau); tau=1 == original SCMK loss
        rr = raw[raw['dataset'] == d]
        if len(rr):
            rr = rr.iloc[0]; rvals = [rr[f'fused_t{t}'] for t in TAUS]
            bi = int(np.argmax(rvals)); rec['rawBW_bestT'] = TAUS[bi]; rec['rawBW_fused'] = rvals[bi]
        # 2x2 factorial: best of the 4 loss cells (fused)
        fr = fac[fac['dataset'].str.startswith(d) | (fac['dataset'] == d)]
        if len(fr):
            fr = fr.iloc[0]
            fcells = {c[:-6]: fr[c] for c in fac.columns if c.endswith('_fused')}
            bk = max(fcells, key=fcells.get); rec['factorial_best'] = bk; rec['factorial_fused'] = fcells[bk]
        # cosine vs kernel: best cosine variant (fused)
        cr = cos[cos['dataset'] == d]
        if len(cr):
            cr = cr.iloc[0]
            ccells = {c[:-6]: cr[c] for c in cos.columns if c.endswith('_fused') and c.startswith('cos')}
            if ccells:
                bk = max(ccells, key=ccells.get); rec['cosine_best'] = bk; rec['cosine_fused'] = ccells[bk]
        rows.append(rec)
    df = pd.DataFrame(rows)
    # mean row (numeric columns only; label/text columns left blank)
    mean = {c: (df[c].mean() if pd.api.types.is_numeric_dtype(df[c]) else '') for c in df.columns}
    mean['dataset'] = 'MEAN'
    df = pd.concat([df, pd.DataFrame([mean])], ignore_index=True)
    return df.round(4)


# ── overview sheet ───────────────────────────────────────────────────────────
def overview_df():
    rows = [
        ['— TWO AXES (do not conflate) —', '', '', '', '', ''],
        ['axis 1: loss version', 'the contrastive TRAINING loss', '', '', '', ''],
        ['axis 2: detector', 'how a trained model -> AUC (raw-RBF / emb-lin / mag / fused / emb-MK)',
         '', '', '', ''],
        ['fused = the ACTUAL SCMK detector', 'max(minmax(emb-lin direction), minmax(mag))', '', '', '', ''],
        ['', '', '', '', '', ''],
        ['Sheet', 'Experiment', 'Config', 'What varies', 'Metric', 'Key finding'],
        ['SCMK_grid_all', 'Original SCMK hyperparameter study (per-seed)',
         '60 datasets, semi split', 'latent_dim{16..256} x lambda{0..1000} x seed{0..4}',
         'SCMK dual-signal fused AUC', 'raw per-run AUC (9000 rows)'],
        ['SCMK_grid_mean', 'Original SCMK, mean over 5 seeds',
         '60 datasets', 'dim x lambda', 'auc_mean +/- auc_std', 'seed-averaged grid'],
        ['SCMK_best', 'Original SCMK best config per dataset',
         '60 datasets', 'best (dim, lambda)', 'best_auc_mean', 'paper SCMK operating point'],
        ['SCMK_param_effect', 'Marginal effect of dim / lambda / seed',
         'over the whole grid', 'dim | lambda | seed', 'mean AUC', 'how params move performance'],
        ['new_temp_sweep_emb', 'NEW: temperature tau, EMBEDDING bandwidth',
         'dim64 lam100 seed2, 19 ds', 'tau in {1..0.05}', 'fused + emb-MK, best tau',
         'per-dataset optimal tau; +0.046 fused vs SCMK (oracle)'],
        ['new_temp_sweep_raw', 'NEW: temperature tau, RAW bandwidth',
         'dim64 lam100 seed2, 19 ds', 'tau in {1..0.05}', 'fused + emb-MK',
         'isolates temperature axis (tau=1 == original SCMK)'],
        ['new_factorial_2x2', 'NEW: bandwidth x temperature factorial',
         'dim64 lam100 seed2, 19 ds', '{raw,emb} bw x {1, 0.2} tau', 'fused / emb-lin / emb-MK',
         'main effects of each change + interaction'],
        ['new_cosine_vs_kernel', 'NEW: cosine NT-Xent vs kernel-logit loss',
         'dim64 lam100 seed2, 19 ds', 'kernel vs cosine tau{0.07,0.1,0.2}', 'fused / emb-lin / emb-MK',
         'does the Gaussian-kernel logit matter vs plain cosine?'],
        ['new_vs_SCMK_summary', 'Head-to-head: every new loss vs original SCMK loss',
         'dim64 lam100 seed2, 19 ds', 'loss architecture', 'fused AUC',
         'best-tau emb-bw is the most consistent win'],
        ['', '', '', '', '', ''],
        ['CAVEAT', 'best_t / best_* are ORACLE (chosen on test AUC) = tuning upper bound.',
         'The deployable number is the best single GLOBAL tau.', '', '', ''],
        ['NOTE', 'The 19-dataset new experiments fix dim=64,lam=100,seed=2; their SCMK baseline',
         '(SCMK_orig_fused / rawBW tau=1 / factorial orig cell / cosine kernel) is that same config,',
         'NOT the multi-seed paper number in SCMK_best.', '', ''],
    ]
    return pd.DataFrame(rows)


# ── write workbook ───────────────────────────────────────────────────────────
def autosize_and_style(ws, freeze='A2', header_row=1, wrap_cols=()):
    hdr_fill = PatternFill('solid', fgColor='DDEBF7')
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = hdr_fill
        cell.alignment = Alignment(vertical='center')
    if freeze:
        ws.freeze_panes = freeze
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 46)
        if letter in wrap_cols:
            for c in col:
                c.alignment = Alignment(wrap_text=True, vertical='top')


def main():
    os.makedirs(RES, exist_ok=True)
    ov = overview_df()
    by_dim, by_lam, by_seed, pivot = param_effect(grid_all)
    summary = build_summary()

    with pd.ExcelWriter(OUT, engine='openpyxl') as xl:
        ov.to_excel(xl, sheet_name='Overview', index=False, header=False)
        summary.to_excel(xl, sheet_name='new_vs_SCMK_summary', index=False)
        annotate_best(emb, 'fused').pipe(lambda d: annotate_best(d, 'embmk')).to_excel(
            xl, sheet_name='new_temp_sweep_emb', index=False)
        annotate_best(raw, 'fused').pipe(lambda d: annotate_best(d, 'embmk')).to_excel(
            xl, sheet_name='new_temp_sweep_raw', index=False)
        fac.to_excel(xl, sheet_name='new_factorial_2x2', index=False)
        cos.to_excel(xl, sheet_name='new_cosine_vs_kernel', index=False)
        # original SCMK study
        by_dim.rename(columns={'mean': 'auc_mean', 'std': 'auc_std', 'count': 'n'}).to_excel(
            xl, sheet_name='SCMK_param_effect', index=False, startrow=1)
        # append lambda / seed / pivot blocks below on the same sheet
        ws = xl.sheets['SCMK_param_effect']
        ws['A1'] = 'Effect of latent_dim (mean over lambda x seed x dataset)'
        r0 = len(by_dim) + 4
        ws.cell(r0, 1, 'Effect of lambda_scatter (mean over dim x seed x dataset)')
        by_lam.rename(columns={'mean': 'auc_mean', 'std': 'auc_std', 'count': 'n'}).to_excel(
            xl, sheet_name='SCMK_param_effect', index=False, startrow=r0)
        r1 = r0 + len(by_lam) + 3
        ws.cell(r1, 1, 'Effect of split_seed (mean over dim x lambda x dataset)')
        by_seed.rename(columns={'mean': 'auc_mean', 'std': 'auc_std', 'count': 'n'}).to_excel(
            xl, sheet_name='SCMK_param_effect', index=False, startrow=r1)
        r2 = r1 + len(by_seed) + 3
        ws.cell(r2, 1, 'dim (rows) x lambda (cols) mean AUC pivot')
        pivot.round(4).to_excel(xl, sheet_name='SCMK_param_effect', startrow=r2)
        grid_best.to_excel(xl, sheet_name='SCMK_best', index=False)
        grid_mean.to_excel(xl, sheet_name='SCMK_grid_mean', index=False)
        grid_all.to_excel(xl, sheet_name='SCMK_grid_all', index=False)

    # styling pass
    import openpyxl
    wb = openpyxl.load_workbook(OUT)
    autosize_and_style(wb['Overview'], freeze=None, header_row=6,
                       wrap_cols=('B', 'C', 'D', 'E', 'F'))
    for sh in ['new_vs_SCMK_summary', 'new_temp_sweep_emb', 'new_temp_sweep_raw',
               'new_factorial_2x2', 'new_cosine_vs_kernel', 'SCMK_best',
               'SCMK_grid_mean', 'SCMK_grid_all']:
        autosize_and_style(wb[sh])
    wb['SCMK_param_effect'].freeze_panes = None
    wb.save(OUT)

    print(f'wrote {OUT}')
    print(f'  sheets: {wb.sheetnames}')
    print(f'  original SCMK grid: {len(grid_all)} runs, {grid_all.dataset.nunique()} datasets, '
          f'dims={sorted(grid_all.latent_dim.unique())}, lambdas={sorted(grid_all.lambda_scatter.unique())}, '
          f'seeds={sorted(grid_all.split_seed.unique())}')
    print('\n  new-loss vs original SCMK (mean fused AUC over 19 datasets):')
    s = summary[summary.dataset != 'MEAN']
    print(f'    original SCMK loss : {s.SCMK_orig_fused.mean():.4f}')
    print(f'    emb-bw best-tau    : {s.embBW_fused.mean():.4f}  ({s.embBW_fused.mean()-s.SCMK_orig_fused.mean():+.4f})')
    print(f'    raw-bw best-tau    : {s.rawBW_fused.mean():.4f}  ({s.rawBW_fused.mean()-s.SCMK_orig_fused.mean():+.4f})')
    print(f'    factorial best cell: {s.factorial_fused.mean():.4f}  ({s.factorial_fused.mean()-s.SCMK_orig_fused.mean():+.4f})')
    print(f'    cosine best        : {s.cosine_fused.mean():.4f}  ({s.cosine_fused.mean()-s.SCMK_orig_fused.mean():+.4f})')


if __name__ == '__main__':
    main()
